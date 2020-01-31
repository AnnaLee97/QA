import os
import subprocess
import datetime, time

import openpyxl
from openpyxl.styles import Font, Alignment


## mem
def sendADB(req) :
    sysMsg = subprocess.getstatusoutput(req)
    return sysMsg[1]

def xlsx_idle_mem(path) :
    if os.path.isfile(path+'mem_info.xlsx') :
        wb = openpyxl.load_workbook(path+'mem_info.xlsx')
    else :
        wb = openpyxl.Workbook()
    
    sh = wb.create_sheet(str(datetime.datetime.now().date()), 0)
    sh.merge_cells('A1:O1')
    sh['A1'] = sendADB('adb shell date')+'   memory information'
    sh.merge_cells('A2:A3')
    sh.merge_cells('B2:F2')
    sh['A2'] = 'Time'
    sh['B2'] = 'RAM info (Kbytes)'
    sh['B3'] = 'TOTAL'
    sh['C3'] = 'Free'
    sh['D3'] = 'Used'
    sh['E3'] = 'Lost'
    sh['F3'] = 'ZRAM'
    sh.merge_cells('G2:N2')
    sh['G2'] = 'Top4 process'
    sh['G3'] = 'Name'
    sh['H3'] = 'Kbytes'
    sh['I3'] = 'Name'
    sh['J3'] = 'Kbytes'
    sh['K3'] = 'Name'
    sh['L3'] = 'Kbytes'
    sh['M3'] = 'Name'
    sh['N3'] = 'Kbytes'
    sh.merge_cells('O2:O3')
    sh['O2'] = 'filename'
    
    # style
    sh.freeze_panes = 'A4'
    sh['A1'].font = Font(size = 20, bold = True)
    rows = sh.max_row
    cols = sh.max_column
    for r in range(1, rows) :
        for c in range(0,cols) :
            sh.cell(row = r+1, column = c+1).alignment = Alignment(horizontal='center', vertical='center')
            sh.cell(row = r+1, column = c+1).font = Font(bold = True)

    wb.save(path+'mem_info.xlsx')


def drop_file_mem(path) :
    date_get = datetime.datetime.now()
    current_time = date_get.strftime('%m%d_%H_%M_%S')
    filename = 'mem_log_'+current_time+'.txt'

    command = 'adb shell dumpsys meminfo > ' + path + 'mem_log/' + filename
    sendADB(command)
    print('   cmd: '+command)

    # time
    tim = sendADB('adb shell date').split(' ')[3]

    # read log file
    obj = []
    f = open(path + 'mem_log/' + filename, 'r')
    temp = f.readlines()
    # time
    obj.append(tim)
    # RAM info
    i = -6
    while i < -1 :
        line = temp[i].split(':')
        tmp = line[1].split('K')
        obj.append(tmp[0])
        i += 1
    # Top4 process
    for i in range(4, 8) :
        line = temp[i].split(':')
        tmp = line[1].split(' ')
        obj.append(tmp[1])
        obj.append(line[0][:-1])
    # filename
    obj.append(filename)

    return obj


def append_xlsx_mem(path, obj) :
    wb = openpyxl.load_workbook(path + 'mem_info.xlsx')
    sh = wb.active
    sh.append(obj)
    wb.save(path + 'mem_info.xlsx')



## cpu
def xlsx_idle_cpu(path) :
    if os.path.isfile(path + 'cpu_info.xlsx') :
        wb = openpyxl.load_workbook(path + 'cpu_info.xlsx')
    else :
        wb = openpyxl.Workbook()
    
    sheet1 = wb.create_sheet(str(datetime.datetime.now().date()), 0)
    #sheet1 = wb.active
    #sheet1.title = 'cpu_info_' + str(datetime.datetime.now().date())
    sheet1.merge_cells('A1:A2')
    sheet1.merge_cells('B1:B2')
    sheet1.merge_cells('C1:H1')
    sheet1['A1'] = 'Date'
    sheet1['B1'] = 'Time'
    sheet1['C1'] = 'CPU usage (%)'
    sheet1['C2'] = 'Total'
    sheet1['D2'] = 'user'
    sheet1['E2'] = 'kernel'
    sheet1['F2'] = 'iowait'
    sheet1['G2'] = 'irq'
    sheet1['H2'] = 'softirq'
    sheet1.merge_cells('I1:J1')
    sheet1['I1'] = 'MAX used pkg'
    sheet1['I2'] = 'pkg name'
    sheet1['J2'] = '%'
    sheet1.merge_cells('K1:K2')
    sheet1['K1'] = 'log file'

    # style
    sheet1.freeze_panes = 'A3'
    rows = sheet1.max_row
    cols = sheet1.max_column
    for r in range(0, rows) :
        for c in range(0,cols) :
            sheet1.cell(row = r+1, column = c+1).alignment = Alignment(horizontal='center', vertical='center')
            sheet1.cell(row = r+1, column = c+1).font = Font(bold = True)

    wb.save(path + 'cpu_info.xlsx')


def drop_file_cpu(path, num, tim) :
    date_get = datetime.datetime.now()
    current_time = date_get.strftime('%m%d_%H_%M_%S')
    filename = 'cpu_log_'+current_time+'.txt'

    #command = 'adb shell top -n 1 -s cpu > D://python/log/' + filename
    command = 'adb shell dumpsys cpuinfo > '+ path + 'cpu_log/' + filename

    # get cpu info txt file
    sendADB(command)
    
    # read txt file
    obj = []
    f = open(path + 'cpu_log/'+filename, 'r')
    temp = f.readlines()
    
    ## date
    info = temp[1].split('(')[1].split(' ')
    if info[1] == tim :
        f.close()
        os.remove(path + 'cpu_log/' + filename)
        return [obj, 1]
    print('   cmd: '+command)
    obj.append(info[0])
    obj.append(info[1])
    ## CPU usage
    line = temp[-1].split(' ')
    obj.append(line[0][:-1]) #total
    tag = ('user', 'kernel', 'iowait', 'irq', 'softirq')
    tag_flag = [0,0,0,0,0]
    j = 0
    for i in range(3, len(line), 3) :
        if i == len(line)-1 :
            tg = line[i][:-1]
        else :
            tg = line[i]
        
        if  j == tag.index(tg) :
            tag_flag[j] = 1
            obj.append(line[i-1][:-1])
            j += 1
        else :
            while j < tag.index(tg) :
                tag_flag[j] = 1
                obj.append('')
                j += 1
            tag_flag[j] = 1
            obj.append(line[i-1][:-1])
            j += 1
    for i in range(j, len(tag_flag)) :
        if tag_flag[j] == 0 :
            obj.append('')
    ## max user
    tmp = temp[2].split(' ')
    obj.append(tmp[3][:-1].split('/')[1]) # max used pkg name
    obj.append(float(tmp[2][:-1])/num) # percentage
    ## log filename
    obj.append(filename)

    f.close()

    return [obj, 0, obj[1]]

def append_xlsx_cpu(obj, path) :
    wb = openpyxl.load_workbook(path + 'cpu_info.xlsx')
    sh = wb.active
    sh.append(obj)

    wb.save(path + 'cpu_info.xlsx')
    
def get_core_num(path) :
    command = 'adb pull /proc/cpuinfo '
    filename = path + 'core.txt'
    sendADB(command+filename)
    
    f = open(filename, 'r')
    temp = f.readlines()
    i = 0
    num = 0
    
    while i < len(temp)-1 :   # doesn't matter what physical id is. so you should modify following codes.
        tmp = temp[i].split(':')
        if tmp[0].find('processor') != -1 :
            num += 1
            for j in range(i+1, len(temp)) :
                if temp[j] == '\n' :
                    i = j+1
                    break
        else :
            break
    f.close()
    os.remove(filename)
    return num





if __name__ == '__main__' :
    print('Choose number below: ')
    print('    1. run UNLimit')
    print('    2. run finite times (User choose)')
    sel = int(input('Write number: '))

    if sel == 2 :
        t = int(input("How many times?: "))
    
    sec = int(input("Every what seconds to check? : "))
    path = os.path.dirname(os.path.abspath(__file__))+'/'
    
    if not os.path.isdir(path+'mem_log') :
        os.mkdir(path + 'mem_log')

    xlsx_idle_mem(path)

    if not os.path.isdir(path+'cpu_log') :
        os.mkdir(path + 'cpu_log')

    xlsx_idle_cpu(path)
    num = get_core_num(path)
    print('     ' + str(num) + ' core used!')
    tim = '0'

    if sel == 2 :
        for i in range(0, t) :
            ## mem
            mem_data = drop_file_mem(path)
            append_xlsx_mem(path, mem_data)

            ## cpu
            cpu_data = drop_file_cpu(path, num, tim)
            if cpu_data[1] == 0 : # if different time
                tim = cpu_data[2]
                append_xlsx_cpu(cpu_data[0], path)

            time.sleep(sec)
    
        print('\nAll done!')
        print('Check  "' + path + 'mem_info.xlsx"')
        print('Check  "'+ path + 'cpu_info.xlsx"\n')

    else :
        while True :
            ## mem
            mem_data = drop_file_mem(path)
            append_xlsx_mem(path, mem_data)

            ## cpu
            cpu_data = drop_file_cpu(path, num, tim)
            if cpu_data[1] == 0 : # if different time
                tim = cpu_data[2]
                append_xlsx_cpu(cpu_data[0], path)

            time.sleep(sec)
