import os
import subprocess
import datetime, time

import openpyxl
from openpyxl.styles import Font, Alignment

###########################################
####  http://github.com/AnnaLee97/QA  ####
###########################################


def sendADB(req) :
    sysMsg = subprocess.getstatusoutput(req)
    print('cmd: '+sysMsg[1])
    return sysMsg[1]

def xlsx_idle(path) :
    if os.path.isfile(path+'mem_info.xlsx') :
        wb = openpyxl.load_workbook(path+'mem_info.xlsx')
    else :
        wb = openpyxl.Workbook()
    
    sh = wb.create_sheet(str(datetime.datetime.now().date()), 0)
    sh.merge_cells('A1:G1')
    sh['A1'] = sendADB('adb shell date')+'   memory information'
    sh.merge_cells('A2:A3')
    sh.merge_cells('B2:F2')
    sh['A2'] = 'Uptime'
    sh['B2'] = 'RAM info (Kbytes)'
    sh['B3'] = 'TOTAL'
    sh['C3'] = 'Free'
    sh['D3'] = 'Used'
    sh['E3'] = 'Lost'
    sh['F3'] = 'ZRAM'
    sh.merge_cells('G2:G3')
    sh['G2'] = 'filename'
    
    # style
    sh.freeze_panes = 'A4'
    sh['A1'].font = Font(size = 20, bold = True)
    rows = sh.max_row
    cols = sh.max_column
    for r in range(1, rows) :
        for c in range(0,cols) :
            sh.cell(row = r+1, column = c+1).alignment = Alignment(horizontal='center', vertical='center')
            sh.cell(row = r+1, column = c+1).font = Font(bold = True)

    wb.save(path+'mem_info.xlsx')


def drop_file(path) :
    date_get = datetime.datetime.now()
    current_time = date_get.strftime('%m%d_%H_%M_%S')
    filename = 'mem_log_'+current_time+'.txt'

    command = 'adb shell dumpsys meminfo > ' + path + 'mem_log/' + filename
    sendADB(command)

    # read log file
    obj = []
    f = open(path + 'mem_log/' + filename, 'r')
    temp = f.readlines()
    # uptime
    tmp = temp[1].split(' ')
    obj.append(tmp[1])
    # RAM info
    i = -6
    while i < -1 :
        line = temp[i].split(':')
        tmp = line[1].split('K')
        obj.append(tmp[0])
        i += 1
    # filename
    obj.append(filename)

    return obj


def append_xlsx(path, obj) :
    wb = openpyxl.load_workbook(path + 'mem_info.xlsx')
    sh = wb.active
    sh.append(obj)
    wb.save(path + 'mem_info.xlsx')









if __name__ == '__main__' :
    t = int(input("How many times?: "))
    sec = int(input("Every what seconds to check? : "))
    path = os.path.dirname(os.path.abspath(__file__))+'/'
    
    if not os.path.isdir(path+'mem_log') :
        os.mkdir(path + 'mem_log')

    xlsx_idle(path)

    for i in range(0, t) :
        data = drop_file(path)
        append_xlsx(path, data)
        time.sleep(sec)
    
    print('\nAll done!')
    print('Check  ' + path + 'mem_info.xlsx\n')
