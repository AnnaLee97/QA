import os
import subprocess
import datetime, time

import openpyxl
from openpyxl.styles import Font, Alignment


############################################
#####  http://github.com/AnnaLee97/QA  #####
############################################




def xlsx_idle(path) :
    if os.path.isfile(path + 'cpu_info.xlsx') :
        wb = openpyxl.load_workbook(path + 'cpu_info.xlsx')
    else :
        wb = openpyxl.Workbook()
    
    sheet1 = wb.create_sheet(str(datetime.datetime.now().date()), 0)
    sheet1.merge_cells('A1:A2')
    sheet1.merge_cells('B1:B2')
    sheet1.merge_cells('C1:H1')
    sheet1['A1'] = 'Date'
    sheet1['B1'] = 'Time'
    sheet1['C1'] = 'CPU usage (%)'
    sheet1['C2'] = 'Total'
    sheet1['D2'] = 'user'
    sheet1['E2'] = 'kernel'
    sheet1['F2'] = 'iowait'
    sheet1['G2'] = 'irq'
    sheet1['H2'] = 'softirq'
    sheet1.merge_cells('I1:J1')
    sheet1['I1'] = 'MAX used pkg'
    sheet1['I2'] = 'pkg name'
    sheet1['J2'] = '%'
    sheet1.merge_cells('K1:K2')
    sheet1['K1'] = 'log file'

    # style
    sheet1.freeze_panes = 'A3'
    rows = sheet1.max_row
    cols = sheet1.max_column
    for r in range(0, rows) :
        for c in range(0,cols) :
            sheet1.cell(row = r+1, column = c+1).alignment = Alignment(horizontal='center', vertical='center')
            sheet1.cell(row = r+1, column = c+1).font = Font(bold = True)

    wb.save(path + 'cpu_info.xlsx')


def sendADB(req) :
    sysMsg = subprocess.getstatusoutput(req)
    return sysMsg[1]

def drop_file(path, num, tim) :
    date_get = datetime.datetime.now()
    current_time = date_get.strftime('%m%d_%H_%M_%S')
    filename = 'cpu_log_'+current_time+'.txt'

    #command = 'adb shell top -n 1 -s cpu > D://python/log/' + filename
    command = 'adb shell dumpsys cpuinfo > '+ path + 'cpu_log/' + filename

    # get cpu info txt file
    sendADB(command)
    
    # read txt file
    obj = []
    f = open(path + 'cpu_log/'+filename, 'r')
    temp = f.readlines()
    
    ## date
    info = temp[1].split('(')[1].split(' ')
    if info[1] == tim :
        f.close()
        os.remove(path + 'cpu_log/' + filename)
        return [obj, 1]
    print('   cmd: '+command)
    obj.append(info[0])
    obj.append(info[1])
    ## CPU usage
    line = temp[-1].split(' ')
    obj.append(line[0][:-1]) #total
    tag = ('user', 'kernel', 'iowait', 'irq', 'softirq')
    tag_flag = [0,0,0,0,0]
    j = 0
    for i in range(3, len(line), 3) :
        if i == len(line)-1 :
            tg = line[i][:-1]
        else :
            tg = line[i]
        
        if  j == tag.index(tg) :
            tag_flag[j] = 1
            obj.append(line[i-1][:-1])
            j += 1
        else :
            while j < tag.index(tg) :
                tag_flag[j] = 1
                obj.append('')
                j += 1
            tag_flag[j] = 1
            obj.append(line[i-1][:-1])
            j += 1
    for i in range(j, len(tag_flag)) :
        if tag_flag[j] == 0 :
            obj.append('')
    ## max user
    tmp = temp[2].split(' ')
    obj.append(tmp[3][:-1].split('/')[1]) # max used pkg name
    obj.append(int(tmp[2][:-1])/num) # percentage
    ## log filename
    obj.append(filename)

    f.close()

    return [obj, 0, obj[1]]

def append_xlsx(obj, path) :
    wb = openpyxl.load_workbook(path + 'cpu_info.xlsx')
    sh = wb.active
    sh.append(obj)

    wb.save(path + 'cpu_info.xlsx')
    
def get_core_num(path) :
    command = 'adb pull /proc/cpuinfo '
    filename = path + 'core.txt'
    sendADB(command+filename)
    
    f = open(filename, 'r')
    temp = f.readlines()
    i = 0
    num = 0
    
    while i < len(temp)-1 :   # doesn't matter what physical id is. so you should modify following codes.
        tmp = temp[i].split(':')
        if tmp[0].find('processor') != -1 :
            num += 1
            for j in range(i+1, len(temp)) :
                if temp[j] == '\n' :
                    i = j+1
                    break
        else :
            break
    f.close()
    os.remove(filename)
    return num
            

            


    




if __name__ == '__main__' :
    t = int(input("How many times?: "))
    sec = int(input("Every what seconds to check? : "))
    path = os.path.dirname(os.path.abspath(__file__))+'/'
    
    if not os.path.isdir(path+'cpu_log') :
        os.mkdir(path + 'cpu_log')

    xlsx_idle(path)
    num = get_core_num(path)
    print('     ' + str(num) + ' core used!')

    tim = '0'
    for i in range(0, t) :
        data = drop_file(path, num, tim)
        if data[1] == 0 : # if different time
            tim = data[2]
            append_xlsx(data[0], path)
        time.sleep(sec)
        
    print('All done!\n')
    print('Check "'+ path + 'cpu_info.xlsx"')
