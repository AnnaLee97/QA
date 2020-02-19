import os
import subprocess
import datetime, time

import openpyxl
from openpyxl.styles import Font, Alignment



def sendADB(req) :
    sysMsg = subprocess.getoutput(req)
    print('    cmd :'+req)
    return sysMsg


def getTemperature(zone) :
    obj = []
    command = 'adb shell cat /sys/class/thermal/'
    # time
    tim = sendADB('adb shell date').split(' ')[3]
    obj.append(tim)
    # info
    for i in zone :
        obj.append(int(sendADB(command + i + '/temp'))/10)
    
    obj.append(zone[obj.index(max(obj[1:]))-1])

    return obj


def idle(path, filename) :
    # thermal zone info
    zone = sendADB('adb shell ls /sys/class/thermal').split('\n')
    # get workbook
    if os.path.isfile(path+filename) :
        wb = openpyxl.load_workbook(path+filename)
    else :
        wb = openpyxl.Workbook()
    # worksheet
    sh = wb.create_sheet(str(datetime.datetime.now().date()), 0)
    sh.merge_cells('A1:A2')
    sh['A1'] = 'Time'
    tmp = 'B1:'+ chr(ord('A')+len(zone))+'1'
    sh.merge_cells(tmp)
    sh['B1'] = "Temperature ('C)"
    for i in range(2, 2+len(zone)) :
        sh.cell(row=2, column=i, value=zone[i-2])
    tmp = chr(ord('A')+len(zone)+1)
    temp = tmp+'1:'+tmp+'2'
    sh.merge_cells(temp)
    sh[tmp+'1'] = "Top temp's zone"
    
    # style
    sh.freeze_panes = 'A3'
    rows = sh.max_row
    cols = sh.max_column
    for r in range(0, rows) :
        for c in range(0,cols) :
            sh.cell(row = r+1, column = c+1).alignment = Alignment(horizontal='center', vertical='center')
            sh.cell(row = r+1, column = c+1).font = Font(bold = True)

    wb.save(path+filename)
    return zone


def append_xlsx(obj, path, filename) :
    wb = openpyxl.load_workbook(path + filename)
    sh = wb.active
    sh.append(obj)

    wb.save(path + filename)
    print('The information has appended!')



if __name__ == '__main__' :
    path = os.path.dirname(os.path.abspath(__file__))+'/'
    date_get = datetime.datetime.now()
    current_time = date_get.strftime('%m%d_%H_%M_%S')
    filename = 'temperature_info_' + current_time + '.xlsx'

    print("How many times to repeat?")
    print("If you want to run this code INFINITE time,")
    print("    then just press ENTER!")
    no = input('Input: ')
    sleep = int(input('Every What seconds to check? : '))
    print('----------------------------------------------------')
    print('Now start!\n')

    zone = idle(path, filename)

    if no == '' or no == '\n' :
        while True :
            append_xlsx(getTemperature(zone), path, filename)
            time.sleep(sleep)
    else :
        for i in range(0, int(no)) :
            append_xlsx(getTemperature(zone), path, filename)
            time.sleep(sleep)
        print('\nAll Done!')
        print('Check '+path+filename)
